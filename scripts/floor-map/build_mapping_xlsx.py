#!/usr/bin/env python3
"""
Таблица для заполнения человеком, знающим здание, — тот же список, что в PDF.

Кому удобнее печатать, а не писать от руки: заполненный файл читается назад
скриптом привязки названий. Колонки «официальное название», «наш номер» и
«замечания» намеренно пустые.

Использование:
    python3 scripts/floor-map/build_mapping_xlsx.py
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FLOORS = [("0", "Подвал"), ("1", "1 этаж"), ("2", "2 этаж"), ("3", "3 этаж")]
MIN_UNNAMED_M2 = 4.0
DISPUTE_PCT = 12.0

HEADERS = [
    ("Этаж / קומה", 14),
    ("№ БТИ / מספר", 12),
    ("Площадь наша, м²", 17),
    ("Площадь на плане", 17),
    ("Официальное название / שם רשמי", 46),
    ("Наш номер / המספר שלנו", 20),
    ("Замечания / הערות", 40),
]


def assign_codes(plan):
    unnamed = [r for r in plan["rooms"]
               if not r["bti_number"] and r["area_computed_m2"] >= MIN_UNNAMED_M2]
    unnamed.sort(key=lambda r: (round(r["centroid_m"][1] / 5), r["centroid_m"][0]))
    return {r["id"]: f"A{i + 1}" for i, r in enumerate(unnamed)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", type=Path, default=Path("data/floor-map"))
    ap.add_argument("--out", type=Path,
                    default=Path("data/floor-map/review/room-mapping-template.xlsx"))
    args = ap.parse_args()
    args.out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Помещения"
    ws.freeze_panes = "A2"

    head_fill = PatternFill("solid", fgColor="E2E8F0")
    warn_fill = PatternFill("solid", fgColor="FFF3D6")
    for col, (title, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = Font(bold=True, size=10)
        c.fill = head_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 32

    row = 2
    total = 0
    for floor, name in FLOORS:
        plan = json.loads((args.dir / f"rooms-floor-{floor}.json").read_text(encoding="utf-8"))
        codes = assign_codes(plan)
        entries = []
        for r in plan["rooms"]:
            if r["bti_number"]:
                entries.append((r["bti_number"], r))
            elif r["id"] in codes:
                entries.append((codes[r["id"]], r))

        def key(t):
            label = t[0]
            if label.startswith("A"):
                return (2, int(label[1:]))
            digits = "".join(ch for ch in label if ch.isdigit())
            return (1, int(digits) if digits else 0)

        for label, r in sorted(entries, key=key):
            printed = r["area_printed_m2"]
            note = ""
            if printed and abs(r["area_computed_m2"] - printed) / printed * 100 > DISPUTE_PCT:
                note = "проверить границы / לבדוק גבולות"
            values = [name, label, r["area_computed_m2"], printed or "", "", "", note]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                if note:
                    cell.fill = warn_fill
            row += 1
            total += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row - 1}"
    wb.save(args.out)
    print(f"{args.out}: строк {total}")


if __name__ == "__main__":
    main()
